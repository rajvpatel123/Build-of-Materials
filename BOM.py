import csv
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook


# ============================================================
# Utility: compare value + unit
# ============================================================

def values_match(v1, u1, v2, u2):
    """Return True if both value and unit strings match (after stripping)."""
    return str(v1).strip() == str(v2).strip() and str(u1).strip() == str(u2).strip()


# ============================================================
# ComponentBox: visual representation of one component
# ============================================================

class ComponentBox:
    def __init__(self, canvas, ref, x, y, angle,
                 comp_type="Unknown", value="", unit="", highlight=None,
                 box_scale=1.0, nc=False):
        """
        canvas    : Tk canvas to draw on
        ref       : reference designator string (e.g. 'L3')
        x, y      : canvas coordinates (already scaled/translated)
        angle     : rotation angle in degrees
        comp_type : 'Capacitor', 'Inductor', etc. (not used visually yet)
        value     : normalized numeric value string (e.g. '2.5')
        unit      : normalized unit string (e.g. 'pF', 'Ohms')
        highlight : None / 'missing' / 'mismatch' (controls box color)
        box_scale : visual scaling factor for box/label size
        nc        : True if this is intentionally marked 'No Component' (N/C)
        """
        self.canvas = canvas
        self.ref = ref
        self.x = x
        self.y = y
        self.angle = float(angle)
        self.comp_type = comp_type
        self.value = value
        self.unit = unit
        self.highlight = highlight
        self.nc = nc

        self.box_scale = box_scale
        base_w = 60
        base_h = 20
        self.width = base_w * self.box_scale
        self.height = base_h * self.box_scale

        self.rect = None
        self.label = None

        self.draw()
        self.bind_events()

    def formatted_label(self):
        """Return the text shown next to the box."""
        base = self.ref
        if self.nc:
            return f"{base} N/C"
        if self.value and self.unit:
            return f"{base} {self.value}{self.unit}"
        elif self.value:
            return f"{base} {self.value}"
        return base

    def draw(self):
        """Draw the rotated rectangle and text label."""
        w = self.width / 2
        h = self.height / 2
        corners = [(-w, -h), (w, -h), (w, h), (-w, h)]

        theta = math.radians(self.angle)
        rotated = []
        for (cx, cy) in corners:
            rx = cx * math.cos(theta) - cy * math.sin(theta)
            ry = cx * math.sin(theta) + cy * math.cos(theta)
            rotated.append((self.x + rx, self.y + ry))

        points = [p for pt in rotated for p in pt]

        if self.highlight == "missing":
            color = "red"
        elif self.highlight == "mismatch":
            color = "yellow"
        else:
            color = "lightblue"

        self.rect = self.canvas.create_polygon(
            points, fill=color, outline="black", width=2
        )

        angle = self.angle % 360
        offset = 6 * self.box_scale
        side_offset = 12 * self.box_scale

        # Place the label on a side that stays readable for common angles
        if 315 <= angle or angle < 45:
            lx, ly = self.x, self.y + h + offset
        elif 45 <= angle < 135:
            lx, ly = self.x, self.y + 30 * self.box_scale + offset
        elif 135 <= angle < 225:
            lx, ly = self.x, self.y - h - offset
        else:
            lx, ly = self.x - w - side_offset, self.y

        self.label = self.canvas.create_text(
            lx, ly, text=self.formatted_label(), font=("Arial", int(9 * self.box_scale))
        )

    def bind_events(self):
        """Attach right-click handler to both box and label."""
        for tag in (self.rect, self.label):
            self.canvas.tag_bind(tag, "<Button-3>", self.right_click)

    def right_click(self, event):
        """Popup editor for value/unit/angle and N/C."""
        popup = tk.Toplevel()
        popup.title(f"Edit {self.ref}")

        tk.Label(popup, text=f"Reference: {self.ref}").pack(pady=5)

        tk.Label(popup, text="Value:").pack()
        val_entry = tk.Entry(popup)
        val_entry.insert(0, self.value)
        val_entry.pack()

        tk.Label(popup, text="Unit:").pack()
        unit_box = ttk.Combobox(
            popup,
            values=["", "pF", "nF", "uF", "pH", "nH", "Ohms"]
        )
        unit_box.set(self.unit)
        unit_box.pack()

        tk.Label(popup, text="Angle:").pack()
        ang_entry = tk.Entry(popup)
        ang_entry.insert(0, str(self.angle))
        ang_entry.pack()

        def save_common(new_value, new_unit, new_angle, nc_flag):
            """Update this box and underlying model, then redraw."""
            self.value = new_value
            self.unit = new_unit
            self.angle = new_angle
            self.nc = nc_flag

            app = getattr(self.canvas, "app_ref", None)
            if app is not None:
                if self.ref in app.xy_data:
                    app.xy_data[self.ref]["value"] = new_value
                    app.xy_data[self.ref]["unit"] = new_unit
                    app.xy_data[self.ref]["angle"] = new_angle
                    app.xy_data[self.ref]["nc"] = nc_flag
                app.redraw()
                app.refresh_production_tree()
            else:
                self.canvas.delete(self.rect)
                self.canvas.delete(self.label)
                self.draw()
                self.bind_events()

        def save():
            """Normal save: keep whatever value/unit the user typed."""
            try:
                new_value = val_entry.get()
                new_unit = unit_box.get()
                new_angle = float(ang_entry.get())
            except ValueError:
                messagebox.showerror("Error", "Angle must be a number.")
                return
            save_common(new_value, new_unit, new_angle, nc_flag=False)
            popup.destroy()

        def set_nc():
            """
            Mark this as No Component:
            - clear value/unit
            - keep angle
            - set nc flag so label shows 'REF N/C' and box goes red
            """
            try:
                new_angle = float(ang_entry.get())
            except ValueError:
                new_angle = self.angle
            save_common("", "", new_angle, nc_flag=True)
            popup.destroy()

        btn_frame = tk.Frame(popup)
        btn_frame.pack(pady=10, fill="x")

        tk.Button(btn_frame, text="Save", command=save).pack(side="left", padx=5)
        tk.Button(btn_frame, text="N/C", command=set_nc).pack(side="left", padx=5)


# ============================================================
# LayoutApp: main GUI/controller
# ============================================================

class LayoutApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Chipset BOM + XY Layout Tool")

        # Data structures
        self.xy_data = {}          # scaled coords + state for drawing
        self.raw_xy_data = {}      # original coords from XY CSV
        self.tuning_boms = []      # list of {ref: {value, unit}}
        self.tuning_bom_names = [] # filenames for display
        self.production_bom = None
        self.production_bom_headers = None

        # Visual scale settings
        self.scale_factor = 100    # XY coordinate scale
        self.box_scale = 1.5       # component box size scale

        # ----------------------------------------------------
        # Menu bar with Scale presets
        # ----------------------------------------------------
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        scale_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Scale", menu=scale_menu)
        scale_menu.add_command(label="Small (0.75x)",
                               command=lambda: self.set_scale(0.75, 80))
        scale_menu.add_command(label="Normal (1.0x)",
                               command=lambda: self.set_scale(1.0, 100))
        scale_menu.add_command(label="Large (1.5x)",
                               command=lambda: self.set_scale(1.5, 140))
        scale_menu.add_command(label="X-Large (2.0x)",
                               command=lambda: self.set_scale(2.0, 180))

        # ----------------------------------------------------
        # Main layout: left canvas, right sidebar
        # ----------------------------------------------------
        main = tk.Frame(root)
        main.pack(fill="both", expand=True)

        # Canvas for component layout
        self.canvas = tk.Canvas(main, bg="white")
        self.canvas.app_ref = self
        self.canvas.pack(side="left", fill="both", expand=True)

        # Sidebar for controls and views
        sidebar = tk.Frame(main)
        sidebar.pack(side="right", fill="y")

        # Top buttons
        top = tk.Frame(sidebar)
        top.pack(fill="x", pady=5)

        tk.Button(top, text="Load XY File",
                  command=self.load_xy_file).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Load Production BOM",
                  command=self.load_production_bom).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Export Production BOM",
                  command=self.export_production_bom).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Load Tuning BOM",
                  command=self.load_tuning_bom_csv).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Save Tuning BOM (Save As)",
                  command=self.save_tuning_bom_csv).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Compare BOMs",
                  command=self.compare_boms).pack(fill="x", padx=5, pady=2)

        # Clear buttons
        tk.Label(sidebar, text="Clear:", font=("Arial", 9, "bold")).pack(anchor="w", padx=5, pady=(8, 0))
        tk.Button(sidebar, text="Clear XY", command=self.clear_xy).pack(fill="x", padx=5, pady=2)
        tk.Button(sidebar, text="Clear Tuning BOMs", command=self.clear_tuning_boms).pack(fill="x", padx=5, pady=2)
        tk.Button(sidebar, text="Clear Production BOM", command=self.clear_production_bom).pack(fill="x", padx=5, pady=2)

        # Tuning BOM list + actions
        tuning_frame = tk.LabelFrame(sidebar, text="Tuning BOMs")
        tuning_frame.pack(fill="both", expand=True, padx=5, pady=4)

        self.tuning_var = tk.IntVar(value=-1)
        self.tuning_buttons_frame = tk.Frame(tuning_frame)
        self.tuning_buttons_frame.pack(fill="both", expand=True)

        tk.Button(tuning_frame, text="Apply Selected",
                  command=self.apply_selected_tuning_bom_sidebar).pack(fill="x", pady=2)
        tk.Button(tuning_frame, text="Compare Tuning BOMs",
                  command=self.compare_tuning_boms).pack(fill="x", pady=2)

        # Production BOM tree view
        prod_frame = tk.LabelFrame(sidebar, text="Production BOM")
        prod_frame.pack(fill="both", expand=True, padx=5, pady=4)

        self.prod_tree = ttk.Treeview(
            prod_frame,
            columns=("ref", "val", "unit"),
            show="headings",
            height=10,
        )
        self.prod_tree.heading("ref", text="Ref")
        self.prod_tree.heading("val", text="Value")
        self.prod_tree.heading("unit", text="Unit")
        self.prod_tree.column("ref", width=70, anchor="w")
        self.prod_tree.column("val", width=80, anchor="w")
        self.prod_tree.column("unit", width=60, anchor="w")
        self.prod_tree.pack(fill="both", expand=True)

        tk.Button(prod_frame, text="Refresh View",
                  command=self.refresh_production_tree).pack(fill="x", pady=2)

    # ========================================================
    # Scale presets
    # ========================================================
    def set_scale(self, box_scale, xy_scale):
        """
        Change visual scale for boxes and XY spacing.
        Recomputes scaled coordinates from stored raw XY data.
        """
        self.box_scale = box_scale
        self.scale_factor = xy_scale

        if self.raw_xy_data:
            for ref, raw in self.raw_xy_data.items():
                if ref not in self.xy_data:
                    self.xy_data[ref] = {
                        "ref": ref,
                        "value": "",
                        "unit": "",
                        "angle": raw.get("angle", 0),
                        "comp_type": self.detect_type(ref),
                        "nc": False,
                    }
                self.xy_data[ref]["x"] = raw["x"] * self.scale_factor
                self.xy_data[ref]["y"] = raw["y"] * self.scale_factor

        self.redraw()

    # ========================================================
    # Clear operations
    # ========================================================
    def clear_xy(self):
        self.xy_data = {}
        self.raw_xy_data = {}
        self.redraw()

    def clear_tuning_boms(self):
        self.tuning_boms = []
        self.tuning_bom_names = []
        self.tuning_var.set(-1)
        for child in self.tuning_buttons_frame.winfo_children():
            child.destroy()
        messagebox.showinfo("Cleared", "All tuning BOMs cleared.")

    def clear_production_bom(self):
        self.production_bom = None
        self.production_bom_headers = None
        self.refresh_production_tree()
        self.redraw()
        messagebox.showinfo("Cleared", "Production BOM cleared.")

    # ========================================================
    # Helpers: auto unit, type detection
    # ========================================================
    def auto_default_unit(self, ref, unit):
        """If a component has a value but no unit, infer one from ref prefix."""
        if unit not in ["", None]:
            return unit

        r = ref.upper()
        if r.startswith("C"):
            return "nF"
        if r.startswith("R"):
            return "Ohms"
        if r.startswith("L"):
            return "nH"
        return ""

    def detect_type(self, ref):
        r = ref.upper()
        if r.startswith("C"):
            return "Capacitor"
        if r.startswith("R"):
            return "Resistor"
        if r.startswith("L"):
            return "Inductor"
        return "Unknown"

    # ========================================================
    # Load XY file (CSV)
    # ========================================================
    def load_xy_file(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not fp:
            return

        data = {}
        raw_data = {}
        try:
            with open(fp, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ref = (
                        row.get("ReferenceID")
                        or row.get("Reference")
                        or row.get("Ref")
                        or row.get("Designator")
                        or ""
                    ).strip()

                    if not ref:
                        continue

                    try:
                        x_raw = float((row.get("X") or "").strip())
                        y_raw = float((row.get("Y") or "").strip())
                        angle = float((row.get("Angle") or "0").strip())
                    except ValueError:
                        continue

                    raw_data[ref] = {
                        "x": x_raw,
                        "y": y_raw,
                        "angle": angle,
                    }

                    x = x_raw * self.scale_factor
                    y = y_raw * self.scale_factor

                    data[ref] = {
                        "ref": ref,
                        "x": x,
                        "y": y,
                        "angle": angle,
                        "value": "",
                        "unit": "",
                        "comp_type": self.detect_type(ref),
                        "nc": False,
                    }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load XY file:\n{e}")
            return

        self.raw_xy_data = raw_data
        self.xy_data = data
        messagebox.showinfo("Loaded", "XY file loaded.")
        self.redraw()

    # ========================================================
    # Load / export production BOM
    # ========================================================
    def load_production_bom(self):
        fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not fp:
            return

        bom, headers = self.parse_production_bom(fp)
        if bom is None:
            return

        self.production_bom = bom
        self.production_bom_headers = headers

        # Apply BOM values to current XY components (except N/C)
        for ref, info in self.xy_data.items():
            if ref in self.production_bom and not info.get("nc", False):
                bval = self.production_bom[ref].get("value", "")
                bunt = self.production_bom[ref].get("unit", "")
                info["value"] = bval
                info["unit"] = bunt

        messagebox.showinfo("Loaded", "Production BOM applied.")
        self.refresh_production_tree()
        self.redraw()

    def refresh_production_tree(self):
        """Populate the right-side tree with the current production BOM."""
        for item in self.prod_tree.get_children():
            self.prod_tree.delete(item)

        if not self.production_bom:
            return

        for ref in sorted(self.production_bom.keys()):
            d = self.production_bom[ref]
            self.prod_tree.insert(
                "",
                "end",
                values=(ref, d.get("value", ""), d.get("unit", "")),
            )

    def export_production_bom(self):
        if not self.production_bom:
            messagebox.showerror("Error", "No production BOM loaded.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Export Production BOM",
        )
        if not save_path:
            return

        wb = Workbook()
        ws = wb.active

        for c, header in enumerate(self.production_bom_headers, start=1):
            ws.cell(row=1, column=c, value=header)

        # Try to preserve original columns
        try:
            col_ref = self.production_bom_headers.index("Reference Designator") + 1
        except ValueError:
            col_ref = 1

        try:
            col_val = self.production_bom_headers.index("Value") + 1
        except ValueError:
            col_val = 2

        try:
            col_unit = self.production_bom_headers.index("Unit") + 1
        except ValueError:
            col_unit = 3

        row = 2
        for ref, d in self.production_bom.items():
            ws.cell(row=row, column=col_ref, value=ref)
            ws.cell(row=row, column=col_val, value=d.get("value", ""))
            ws.cell(row=row, column=col_unit, value=d.get("unit", ""))
            row += 1

        wb.save(save_path)
        messagebox.showinfo("Saved", f"Production BOM exported:\n{save_path}")

    # ========================================================
    # Parse production BOM (handles resistors and 0 Ohm)
    # ========================================================
    def parse_production_bom(self, filepath):
        """
        Parser for Excel BOM that:

        - Detects header row with 'Reference Designator' and 'Value'
        - Uses 'Reference Designator' for refs
        - Uses 'Value' for value/unit parsing
        - Uses 'Type' column (C) to identify resistors ('Res')
        - Splits multi-reference cells like 'R4, R5'
        - Keeps 0 Ohm resistors as numeric 0, unit Ohms
        - Treats numeric 0 as missing for non-resistor rows
        """
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open file:\n{e}")
            return None, None

        REF_KEYS = ["reference designator"]
        VALUE_KEYS = ["value"]

        header_row = None
        col_ref = col_val = None

        # Find header row
        for r in range(1, 60):
            row_vals = [
                (c, str(ws.cell(r, c).value).strip().lower())
                for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value
            ]
            possible_ref = [c for c, t in row_vals if any(k in t for k in REF_KEYS)]
            possible_val = [c for c, t in row_vals if any(k in t for k in VALUE_KEYS)]
            if possible_ref and possible_val:
                header_row = r
                col_ref = possible_ref[0]
                col_val = possible_val[0]
                break

        if not header_row:
            messagebox.showerror("Error", "Could not find BOM header row.")
            return None, None

        headers = [
            ws.cell(header_row, c).value or ""
            for c in range(1, ws.max_column + 1)
        ]

        bom = {}

        # Walk all data rows
        for r in range(header_row + 1, ws.max_row + 1):
            ref_cell = ws.cell(r, col_ref).value
            val_cell = ws.cell(r, col_val).value

            if not ref_cell:
                # Skip blank reference rows but continue to later rows
                continue

            # Split cell like "R4, R5" into separate refs
            refs = (
                str(ref_cell)
                .replace(";", ",")
                .replace("/", ",")
                .replace(" ", ",")
                .split(",")
            )

            raw_val_str = str(val_cell).strip() if val_cell else ""

            numeric = self.extract_numeric(raw_val_str)
            inferred_unit = self.extract_unit(raw_val_str, explicit_unit="")

            # Type column C: "Cap", "Ind", "Res", etc.
            type_cell = ws.cell(r, 3).value
            type_str = str(type_cell or "").strip().lower()
            is_resistor_row = "res" in type_str

            # Special case: keep 0 Ohm resistors as real parts
            is_zero_ohm = (
                is_resistor_row and
                numeric == "0" and
                inferred_unit == "Ohms"
            )

            if numeric == "0" and not is_zero_ohm:
                numeric = ""
                unit = ""
            else:
                unit = inferred_unit
                if refs:
                    unit = self.auto_default_unit(refs[0], unit)

            for ref in refs:
                ref = ref.strip()
                if not ref:
                    continue
                bom[ref] = {
                    "value": numeric,
                    "unit": unit,
                }

        return bom, headers

    # ========================================================
    # Numeric + unit extraction helpers
    # ========================================================
    def extract_numeric(self, raw):
        if raw in ["", None]:
            return ""
        s = str(raw).lower().replace(" ", "")
        out = ""
        for ch in s:
            if ch.isdigit() or ch == ".":
                out += ch
        return out

    def extract_unit(self, raw, explicit_unit):
        if explicit_unit not in ["", None]:
            return str(explicit_unit).strip()

        if raw in ["", None]:
            return ""

        s = str(raw).lower()

        if "ohm" in s or s.endswith("r"):
            return "Ohms"
        if "pf" in s:
            return "pF"
        if "nf" in s:
            return "nF"
        if "uf" in s or "μf" in s:
            return "uF"
        if "nh" in s:
            return "nH"
        if "uh" in s or "μh" in s:
            return "uH"

        return ""

    # ========================================================
    # Load / save tuning BOM (CSV)
    # ========================================================
    def load_tuning_bom_csv(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not fp:
            return

        tuning = {}
        try:
            with open(fp, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ref = (
                        row.get("ReferenceID")
                        or row.get("Reference")
                        or row.get("Ref")
                        or ""
                    ).strip()
                    if not ref:
                        continue

                    raw_val = (row.get("Value") or "").strip()
                    raw_unit = (row.get("Unit") or "").strip()

                    parts = raw_val.split()
                    raw_val = parts[-1] if parts else ""

                    numeric = self.extract_numeric(raw_val)
                    unit = self.extract_unit(raw_val, explicit_unit=raw_unit)

                    # Treat 0 as "no component" in tuning files
                    if numeric == "0":
                        numeric = ""
                        unit = ""
                    else:
                        unit = self.auto_default_unit(ref, unit)

                    tuning[ref] = {
                        "value": numeric,
                        "unit": unit,
                    }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load tuning BOM:\n{e}")
            return

        self.tuning_boms.append(tuning)
        self.tuning_bom_names.append(fp)

        # Add radio button entry in sidebar
        idx = len(self.tuning_boms) - 1
        rb = tk.Radiobutton(
            self.tuning_buttons_frame,
            text=f"Tuning {idx + 1}",
            variable=self.tuning_var,
            value=idx,
            anchor="w",
            justify="left"
        )
        rb.pack(fill="x", padx=2, pady=1)
        if self.tuning_var.get() == -1:
            self.tuning_var.set(0)

        messagebox.showinfo("Loaded", f"Tuning BOM loaded:\n{fp}")

    def save_tuning_bom_csv(self):
        if not self.xy_data:
            messagebox.showerror("Error", "Load XY file first.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            title="Save Tuning BOM As",
        )
        if not save_path:
            return

        with open(save_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ReferenceID", "X", "Y", "Angle", "Value", "Unit"])

            for ref, info in self.xy_data.items():
                if ref in self.raw_xy_data:
                    x_raw = self.raw_xy_data[ref]["x"]
                    y_raw = self.raw_xy_data[ref]["y"]
                else:
                    x_raw = info["x"] / self.scale_factor
                    y_raw = info["y"] / self.scale_factor

                writer.writerow([
                    ref,
                    x_raw,
                    y_raw,
                    info["angle"],
                    info.get("value", ""),
                    info.get("unit", ""),
                ])

        messagebox.showinfo("Saved", f"Tuning BOM saved:\n{save_path}")

    # ========================================================
    # Apply selected tuning BOM to layout
    # ========================================================
    def apply_selected_tuning_bom_sidebar(self):
        idx = self.tuning_var.get()
        if idx < 0 or idx >= len(self.tuning_boms):
            messagebox.showerror("Error", "No tuning BOM selected.")
            return

        bom = self.tuning_boms[idx]
        for ref, info in self.xy_data.items():
            if info.get("nc", False):
                # Preserve explicit N/C overrides
                continue
            if ref in bom:
                info["value"] = bom[ref]["value"]
                info["unit"] = bom[ref]["unit"]
        self.redraw()

    # ========================================================
    # Compare tuning BOM vs production BOM
    # ========================================================
    def compare_boms(self):
        if not self.production_bom:
            messagebox.showerror("Error", "Load a production BOM first.")
            return
        if not self.tuning_boms:
            messagebox.showerror("Error", "Load at least one tuning BOM.")
            return

        win = tk.Toplevel()
        win.title("Compare Tuning vs Production")

        tk.Label(win, text="Select tuning BOM:").pack()
        names = [
            f"Tuning {i + 1}: {self.tuning_bom_names[i]}"
            for i in range(len(self.tuning_boms))
        ]
        sel = tk.StringVar()
        combo = ttk.Combobox(win, textvariable=sel, values=names, width=60)
        combo.pack()
        combo.current(0)

        def do_compare():
            tuning = self.tuning_boms[combo.current()]
            win.destroy()
            self.show_bom_vs_production(tuning)

        tk.Button(win, text="Compare", command=do_compare).pack(pady=10)

    def show_bom_vs_production(self, tuning_bom):
        """Show only refs whose value/unit differ between tuning and production."""
        win = tk.Toplevel()
        win.title("Tuning vs Production Differences")

        tree = ttk.Treeview(
            win,
            columns=("ref", "prod_val", "prod_unit", "tun_val", "tun_unit"),
            show="headings",
        )
        tree.pack(fill="both", expand=True)

        tree.heading("ref", text="Ref")
        tree.heading("prod_val", text="Prod Val")
        tree.heading("prod_unit", text="Prod Unit")
        tree.heading("tun_val", text="Tuning Val")
        tree.heading("tun_unit", text="Tuning Unit")

        refs = sorted(set(self.production_bom.keys()) | set(tuning_bom.keys()))
        for ref in refs:
            p = self.production_bom.get(ref, {})
            t = tuning_bom.get(ref, {})
            p_val = p.get("value", "")
            p_unit = p.get("unit", "")
            t_val = t.get("value", "")
            t_unit = t.get("unit", "")

            # Only interested in tuning entries that exist
            if not (t_val or t_unit):
                continue

            # Only show non-matching value+unit rows
            if not values_match(p_val, p_unit, t_val, t_unit):
                tree.insert(
                    "",
                    "end",
                    values=(ref, p_val, p_unit, t_val, t_unit),
                )

    # ========================================================
    # Compare tuning BOM vs tuning BOM
    # ========================================================
    def compare_tuning_boms(self):
        if len(self.tuning_boms) < 2:
            messagebox.showerror("Error", "Load at least two tuning BOMs.")
            return

        win = tk.Toplevel()
        win.title("Compare Tuning BOMs")

        names = [
            f"Tuning {i + 1}: {self.tuning_bom_names[i]}"
            for i in range(len(self.tuning_boms))
        ]

        tk.Label(win, text="Select BOM A:").pack()
        selA = tk.StringVar()
        comboA = ttk.Combobox(win, textvariable=selA, values=names, width=60)
        comboA.pack()
        comboA.current(0)

        tk.Label(win, text="Select BOM B:").pack()
        selB = tk.StringVar()
        comboB = ttk.Combobox(win, textvariable=selB, values=names, width=60)
        comboB.pack()
        comboB.current(1)

        def do_compare():
            bomA = self.tuning_boms[comboA.current()]
            bomB = self.tuning_boms[comboB.current()]
            win.destroy()
            self.show_tuning_difference_table(bomA, bomB)

        tk.Button(win, text="Compare", command=do_compare).pack(pady=10)

    def show_tuning_difference_table(self, bomA, bomB):
        """Show only refs where tuning BOM A and B disagree (value+unit)."""
        win = tk.Toplevel()
        win.title("Tuning BOM Differences")

        tree = ttk.Treeview(
            win,
            columns=("ref", "A_val", "A_unit", "B_val", "B_unit"),
            show="headings",
        )
        tree.pack(fill="both", expand=True)

        tree.heading("ref", text="Ref")
        tree.heading("A_val", text="A Value")
        tree.heading("A_unit", text="A Unit")
        tree.heading("B_val", text="B Value")
        tree.heading("B_unit", text="B Unit")

        refs = sorted(set(bomA.keys()) | set(bomB.keys()))
        for ref in refs:
            A_val = bomA.get(ref, {}).get("value", "")
            A_unit = bomA.get(ref, {}).get("unit", "")
            B_val = bomB.get(ref, {}).get("value", "")
            B_unit = bomB.get(ref, {}).get("unit", "")

            # Only show non-matching rows
            if not values_match(A_val, A_unit, B_val, B_unit):
                tree.insert(
                    "",
                    "end",
                    values=(ref, A_val, A_unit, B_val, B_unit),
                )

    # ========================================================
    # Redraw canvas (auto-centered)
    # ========================================================
    def redraw(self):
        """Clear and redraw all components based on xy_data."""
        self.canvas.delete("all")

        if not self.xy_data:
            return

        xs = [info["x"] for info in self.xy_data.values()]
        ys = [info["y"] for info in self.xy_data.values()]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)

        canvas_w = int(self.canvas.winfo_width() or 1500)
        canvas_h = int(self.canvas.winfo_height() or 900)

        width_span = max_x - min_x if max_x != min_x else 1
        height_span = max_y - min_y if max_y != min_y else 1

        offset_x = (canvas_w - width_span) / 2 - min_x
        offset_y = (canvas_h - height_span) / 2 - min_y

        for ref, info in self.xy_data.items():
            val = info.get("value", "")
            unit = info.get("unit", "")
            angle = info.get("angle", 0)
            nc_flag = info.get("nc", False)

            # N/C parts have empty value/unit and are always red (missing)
            if nc_flag:
                val = ""
                unit = ""

            if val and not unit:
                unit = self.auto_default_unit(ref, unit)
                info["unit"] = unit

            highlight = None

            is_zero_ohm = (str(val).strip() == "0") and (str(unit).strip() == "Ohms")

            # Missing value -> red (except 0 Ohm)
            if (val in ["", None]) and not is_zero_ohm:
                highlight = "missing"
            else:
                # Only check production mismatch when not marked N/C
                if self.production_bom and ref in self.production_bom and not nc_flag:
                    pval = self.production_bom[ref].get("value", "")
                    punit = self.production_bom[ref].get("unit", "")
                    if pval or punit:
                        if not values_match(val, unit, pval, punit):
                            highlight = "mismatch"

            draw_x = info["x"] + offset_x
            draw_y = info["y"] + offset_y

            ComponentBox(
                self.canvas,
                ref,
                draw_x,
                draw_y,
                angle,
                comp_type=info.get("comp_type", "Unknown"),
                value=val,
                unit=unit,
                highlight=highlight,
                box_scale=self.box_scale,
                nc=nc_flag,
            )


# ============================================================
# Application entry point
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    root.state("zoomed")
    app = LayoutApp(root)
    root.mainloop()

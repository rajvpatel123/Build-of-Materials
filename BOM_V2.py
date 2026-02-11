import os
import csv
import math
import re
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook, Workbook


# ============================================================
# Utility
# ============================================================

def values_match(v1, u1, v2, u2):
    return str(v1).strip() == str(v2).strip() and str(u1).strip() == str(u2).strip()


# ============================================================
# Board model: XY + versions V0,V1,V2...
# ============================================================

class Board:
    """
    Single board file:
      - XY coords (ref, x, y, angle)
      - multiple version columns (V0, V1, V2, ...) for tuning history
    """

    def __init__(self, path):
        self.path = path
        self.name = os.path.basename(path)
        self.xy = {}          # ref -> {x, y, angle}
        self.versions = []    # list of dicts: [{ref: {value, unit}}, ...]
        self.version_names = []  # e.g. ["V0", "V1", ...]
        self.version_meta = []  # list of {timestamp, notes}

    @classmethod
    def load_from_csv(cls, path):
        board = cls(path)
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return board

            version_cols = [c for c in reader.fieldnames if c and c.upper().startswith("V")]
            board.version_names = version_cols[:]

            for row in reader:
                ref = (row.get("ReferenceID") or row.get("Ref") or "").strip()
                if not ref:
                    continue
                try:
                    x = float((row.get("X") or "").strip())
                    y = float((row.get("Y") or "").strip())
                    angle = float((row.get("Angle") or "0").strip())
                except ValueError:
                    continue

                board.xy[ref] = {"x": x, "y": y, "angle": angle}

                while len(board.versions) < len(version_cols):
                    board.versions.append({})

                for idx, col in enumerate(version_cols):
                    raw = (row.get(col) or "").strip()
                    if not raw:
                        continue
                    vdict = board.versions[idx]
                    vdict.setdefault(ref, {"value": "", "unit": ""})
                    vdict[ref]["value"] = raw
                    vdict[ref]["unit"] = ""

        return board

    @staticmethod
    def _normalize_unit(raw_unit):
        unit_map = {
            "pf": "pF",
            "nf": "nF",
            "uf": "uF",
            "ohms": "Ohms",
            "ohm": "Ohms",
            "r": "Ohms",
            "nh": "nH",
            "uh": "uH",
        }
        key = str(raw_unit or "").strip().lower()
        return unit_map.get(key, raw_unit)

    @classmethod
    def _split_value_unit(cls, raw):
        if raw in ["", None]:
            return "", ""
        s = str(raw).strip()
        if not s:
            return "", ""
        match = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*([a-zA-Zμ]*)\s*$", s)
        if match:
            value = match.group(1)
            unit = cls._normalize_unit(match.group(2))
            return value, unit
        value = "".join(ch for ch in s if ch.isdigit() or ch == ".")
        unit = s.replace(value, "").strip()
        unit = cls._normalize_unit(unit)
        return value, unit

    @staticmethod
    def _join_value_unit(value, unit):
        value = str(value or "").strip()
        unit = str(unit or "").strip()
        if not value:
            return ""
        if unit:
            return f"{value}{unit}"
        return value

    @classmethod
    def load_from_xlsx(cls, path):
        board = cls(path)
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        headers = [
            (idx + 1, str(ws.cell(1, idx + 1).value or "").strip())
            for idx in range(ws.max_column)
        ]
        header_map = {name.lower(): col for col, name in headers if name}

        ref_col = header_map.get("referenceid") or header_map.get("reference id") or header_map.get("ref") or 1
        x_col = header_map.get("x") or 2
        y_col = header_map.get("y") or 3
        angle_col = header_map.get("angle") or 4

        version_columns = {}
        for col, name in headers:
            if not name:
                continue
            base_match = re.match(r"^(v\d+)(?:_(timestamp|notes))?$", name.strip(), re.IGNORECASE)
            if not base_match:
                continue
            base = base_match.group(1).upper()
            suffix = (base_match.group(2) or "value").lower()
            version_columns.setdefault(base, {"value": None, "timestamp": None, "notes": None})
            version_columns[base][suffix] = col

        ordered_versions = []
        for col, name in headers:
            if not name:
                continue
            if re.match(r"^v\d+$", name.strip(), re.IGNORECASE):
                ordered_versions.append(name.strip().upper())

        board.version_names = ordered_versions[:]
        board.versions = [{} for _ in board.version_names]
        board.version_meta = [{"timestamp": "", "notes": ""} for _ in board.version_names]

        meta_found = [False for _ in board.version_names]
        for row in range(2, ws.max_row + 1):
            ref_cell = ws.cell(row, ref_col).value
            if not ref_cell:
                continue
            ref = str(ref_cell).strip()
            try:
                x = float(ws.cell(row, x_col).value)
                y = float(ws.cell(row, y_col).value)
            except (TypeError, ValueError):
                continue
            angle_raw = ws.cell(row, angle_col).value
            try:
                angle = float(angle_raw) if angle_raw is not None else 0.0
            except (TypeError, ValueError):
                angle = 0.0

            board.xy[ref] = {"x": x, "y": y, "angle": angle}

            for idx, ver_name in enumerate(board.version_names):
                cols = version_columns.get(ver_name, {})
                value_col = cols.get("value")
                if value_col:
                    raw_val = ws.cell(row, value_col).value
                    if raw_val not in ["", None]:
                        value, unit = cls._split_value_unit(raw_val)
                        board.versions[idx].setdefault(ref, {"value": "", "unit": ""})
                        board.versions[idx][ref]["value"] = value
                        board.versions[idx][ref]["unit"] = unit

                if not meta_found[idx]:
                    ts_col = cols.get("timestamp")
                    notes_col = cols.get("notes")
                    ts_val = ws.cell(row, ts_col).value if ts_col else ""
                    notes_val = ws.cell(row, notes_col).value if notes_col else ""
                    if ts_val or notes_val:
                        board.version_meta[idx]["timestamp"] = str(ts_val or "")
                        board.version_meta[idx]["notes"] = str(notes_val or "")
                        meta_found[idx] = True

        return board

    def append_version_from_xy(self, xy_data, timestamp, notes):
        """
        Create a new version Vn from current xy_data values.
        Only refs with non-empty value/unit are stored.
        """
        bom = {}
        for ref, info in xy_data.items():
            if info.get("nc", False):
                continue
            val = str(info.get("value", "")).strip()
            unit = str(info.get("unit", "")).strip()
            if not val and not unit:
                continue
            bom[ref] = {"value": val, "unit": unit}
        self.versions.append(bom)
        self.version_names.append(f"V{len(self.versions) - 1}")
        self.version_meta.append({"timestamp": timestamp, "notes": notes})

    def save_to_csv(self):
        """
        Write board to CSV with V0..Vn columns.
        Value/unit are stored as plain value strings per version.
        """
        if not self.xy:
            return

        if not self.version_names:
            self.version_names = [f"V{i}" for i in range(len(self.versions))]

        fieldnames = ["ReferenceID", "X", "Y", "Angle"] + self.version_names
        refs = sorted(self.xy.keys())

        with open(self.path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for ref in refs:
                row = {
                    "ReferenceID": ref,
                    "X": self.xy[ref]["x"],
                    "Y": self.xy[ref]["y"],
                    "Angle": self.xy[ref]["angle"],
                }
                for idx, col in enumerate(self.version_names):
                    val = ""
                    if idx < len(self.versions):
                        v = self.versions[idx].get(ref, {})
                        val = v.get("value", "")
                    row[col] = val
                writer.writerow(row)

    def save_to_xlsx(self):
        if not self.xy:
            return

        if not self.version_names:
            self.version_names = [f"V{i}" for i in range(len(self.versions))]

        wb = Workbook()
        ws = wb.active

        headers = ["ReferenceID", "X", "Y", "Angle"]
        for idx, name in enumerate(self.version_names):
            headers.extend([
                name,
                f"{name}_Timestamp",
                f"{name}_Notes",
            ])

        ws.append(headers)
        refs = sorted(self.xy.keys())
        for ref in refs:
            row = [
                ref,
                self.xy[ref]["x"],
                self.xy[ref]["y"],
                self.xy[ref]["angle"],
            ]
            for idx, name in enumerate(self.version_names):
                version = self.versions[idx] if idx < len(self.versions) else {}
                val = ""
                if ref in version:
                    val = self._join_value_unit(version[ref].get("value", ""), version[ref].get("unit", ""))
                meta = self.version_meta[idx] if idx < len(self.version_meta) else {"timestamp": "", "notes": ""}
                row.extend([val, meta.get("timestamp", ""), meta.get("notes", "")])
            ws.append(row)

        wb.save(self.path)


# ============================================================
# Production BOM model (with tolerance/size/rating)
# ============================================================

class ProductionBOM:
    def __init__(self):
        self.entries = {}   # ref -> {value, unit, tolerance, size_eia, rating}
        self.headers = []

    @classmethod
    def from_xlsx(cls, path, auto_unit_fn):
        inst = cls()
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        REF_KEYS = ["reference designator"]
        VALUE_KEYS = ["value"]
        TOL_KEYS = ["tolerance"]
        SIZE_KEYS = ["size (eia)"]
        RATE_KEYS = ["rating"]

        header_row = None
        col_ref = col_val = None
        col_tol = col_size = col_rate = None

        for r in range(1, 60):
            row_vals = [
                (c, str(ws.cell(r, c).value).strip().lower())
                for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value
            ]
            possible_ref = [c for c, t in row_vals if any(k in t for k in REF_KEYS)]
            possible_val = [c for c, t in row_vals if any(k in t for k in VALUE_KEYS)]
            if possible_ref and possible_val:
                header_row = r
                col_ref = possible_ref[0]
                col_val = possible_val[0]
                for c, t in row_vals:
                    if any(k in t for k in TOL_KEYS):
                        col_tol = c
                    if any(k in t for k in SIZE_KEYS):
                        col_size = c
                    if any(k in t for k in RATE_KEYS):
                        col_rate = c
                break

        if not header_row:
            raise ValueError("Could not find BOM header row")

        inst.headers = [
            ws.cell(header_row, c).value or ""
            for c in range(1, ws.max_column + 1)
        ]

        for r in range(header_row + 1, ws.max_row + 1):
            ref_cell = ws.cell(r, col_ref).value
            val_cell = ws.cell(r, col_val).value
            tol_cell = ws.cell(r, col_tol).value if col_tol else ""
            size_cell = ws.cell(r, col_size).value if col_size else ""
            rate_cell = ws.cell(r, col_rate).value if col_rate else ""

            if not ref_cell:
                continue

            refs = (
                str(ref_cell)
                .replace(";", ",")
                .replace("/", ",")
                .replace(" ", ",")
                .split(",")
            )

            raw_val = str(val_cell).strip() if val_cell else ""
            numeric = "".join(ch for ch in raw_val if ch.isdigit() or ch == ".")
            unit = ""

            if numeric == "0":
                numeric = ""
                unit = ""
            else:
                if refs:
                    unit = auto_unit_fn(refs[0], unit)

            tol = str(tol_cell).strip() if tol_cell else ""
            size_eia = str(size_cell).strip() if size_cell else ""
            rating = str(rate_cell).strip() if rate_cell else ""

            for ref in refs:
                ref = ref.strip()
                if not ref:
                    continue
                inst.entries[ref] = {
                    "value": numeric,
                    "unit": unit,
                    "tolerance": tol,
                    "size_eia": size_eia,
                    "rating": rating,
                }

        return inst


# ============================================================
# Component box widget (with left + right click)
# ============================================================

class ComponentBox:
    def __init__(self, canvas, ref, x, y, angle,
                 comp_type="Unknown", value="", unit="", highlight=None,
                 box_scale=1.5, nc=False, on_left_click=None):
        self.canvas = canvas
        self.ref = ref
        self.x = x
        self.y = y
        self.angle = float(angle)
        self.comp_type = comp_type
        self.value = value
        self.unit = unit
        self.highlight = highlight
        self.nc = nc
        self.on_left_click = on_left_click

        self.box_scale = box_scale
        base_w = 60
        base_h = 20
        self.width = base_w * self.box_scale
        self.height = base_h * self.box_scale

        self.rect = None
        self.label = None

        self.draw()
        self.bind_events()

    def formatted_label(self):
        base = self.ref
        if self.nc:
            return f"{base} N/C"
        if self.value and self.unit:
            return f"{base} {self.value}{self.unit}"
        elif self.value:
            return f"{base} {self.value}"
        return base

    def draw(self):
        w = self.width / 2
        h = self.height / 2
        corners = [(-w, -h), (w, -h), (w, h), (-w, h)]

        theta = math.radians(self.angle)
        rotated = []
        for (cx, cy) in corners:
            rx = cx * math.cos(theta) - cy * math.sin(theta)
            ry = cx * math.sin(theta) + cy * math.cos(theta)
            rotated.append((self.x + rx, self.y + ry))

        points = [p for pt in rotated for p in pt]

        if self.highlight == "missing":
            color = "red"
        elif self.highlight == "mismatch":
            color = "yellow"
        else:
            color = "lightblue"

        self.rect = self.canvas.create_polygon(
            points, fill=color, outline="black", width=2
        )

        angle = self.angle % 360
        offset = 6 * self.box_scale
        side_offset = 12 * self.box_scale

        if 315 <= angle or angle < 45:
            lx, ly = self.x, self.y + h + offset
        elif 45 <= angle < 135:
            lx, ly = self.x, self.y + 30 * self.box_scale + offset
        elif 135 <= angle < 225:
            lx, ly = self.x, self.y - h - offset
        else:
            lx, ly = self.x - w - side_offset, self.y

        self.label = self.canvas.create_text(
            lx, ly, text=self.formatted_label(), font=("Arial", int(9 * self.box_scale))
        )

    def bind_events(self):
        for tag in (self.rect, self.label):
            self.canvas.tag_bind(tag, "<Button-1>", self.left_click)
            self.canvas.tag_bind(tag, "<Button-3>", self.right_click)

    def left_click(self, event):
        if self.on_left_click:
            self.on_left_click(self.ref)

    def right_click(self, event):
        """Popup editor for value/unit/angle and N/C."""
        popup = tk.Toplevel()
        popup.title(f"Edit {self.ref}")

        tk.Label(popup, text=f"Reference: {self.ref}").pack(pady=5)

        tk.Label(popup, text="Value:").pack()
        val_entry = tk.Entry(popup)
        val_entry.insert(0, self.value)
        val_entry.pack()

        tk.Label(popup, text="Unit:").pack()
        unit_box = ttk.Combobox(
            popup,
            values=["", "pF", "nF", "uF", "pH", "nH", "Ohms"]
        )
        unit_box.set(self.unit)
        unit_box.pack()

        tk.Label(popup, text="Angle:").pack()
        ang_entry = tk.Entry(popup)
        ang_entry.insert(0, str(self.angle))
        ang_entry.pack()

        def save_common(new_value, new_unit, new_angle, nc_flag):
            self.value = new_value
            self.unit = new_unit
            self.angle = new_angle
            self.nc = nc_flag

            app = getattr(self.canvas, "app_ref", None)
            if app is not None and hasattr(app, "xy_data"):
                if self.ref in app.xy_data:
                    app.xy_data[self.ref]["value"] = new_value
                    app.xy_data[self.ref]["unit"] = new_unit
                    app.xy_data[self.ref]["angle"] = new_angle
                    app.xy_data[self.ref]["nc"] = nc_flag
                app.redraw()

        def save():
            try:
                new_value = val_entry.get()
                new_unit = unit_box.get()
                new_angle = float(ang_entry.get())
            except ValueError:
                messagebox.showerror("Error", "Angle must be a number.")
                return
            save_common(new_value, new_unit, new_angle, nc_flag=False)
            popup.destroy()

        def set_nc():
            try:
                new_angle = float(ang_entry.get())
            except ValueError:
                new_angle = self.angle
            save_common("", "", new_angle, nc_flag=True)
            popup.destroy()

        btn_frame = tk.Frame(popup)
        btn_frame.pack(pady=10, fill="x")

        tk.Button(btn_frame, text="Save", command=save).pack(side="left", padx=5)
        tk.Button(btn_frame, text="N/C", command=set_nc).pack(side="left", padx=5)


# ============================================================
# Main application
# ============================================================

class LayoutAppV2:
    def __init__(self, root):
        self.root = root
        self.root.title("Chipset BOM + XY Layout Tool v2")

        self.xy_data = {}
        self.raw_xy_data = {}

        self.current_board = None
        self.current_board_version = None

        self.production_bom = None

        self.scale_factor = 100
        self.box_scale = 1.5

        self.detail_tol_var = tk.StringVar(value="")
        self.detail_size_var = tk.StringVar(value="")
        self.detail_rate_var = tk.StringVar(value="")

        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        scale_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Scale", menu=scale_menu)
        scale_menu.add_command(label="Small (0.75x)",
                               command=lambda: self.set_scale(0.75, 80))
        scale_menu.add_command(label="Normal (1.0x)",
                               command=lambda: self.set_scale(1.0, 100))
        scale_menu.add_command(label="Large (1.5x)",
                               command=lambda: self.set_scale(1.5, 140))
        scale_menu.add_command(label="X-Large (2.0x)",
                               command=lambda: self.set_scale(2.0, 180))

        main = tk.Frame(root)
        main.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(main, bg="white")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.app_ref = self

        sidebar = tk.Frame(main)
        sidebar.pack(side="right", fill="y")

        top = tk.Frame(sidebar)
        top.pack(fill="x", pady=5)

        tk.Button(top, text="Load XY File",
                  command=self.load_xy_file).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Load Production BOM",
                  command=self.load_production_bom).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Export Production BOM",
                  command=self.export_production_bom).pack(fill="x", padx=5, pady=2)

        tk.Button(top, text="Create Board",
                  command=self.create_board).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Load Board (XLSX)",
                  command=self.load_board_xlsx).pack(fill="x", padx=5, pady=2)
        tk.Button(top, text="Save New Board Version",
                  command=self.save_new_board_version).pack(fill="x", padx=5, pady=2)

        tk.Button(top, text="Compare Board Versions",
                  command=self.compare_board_versions).pack(fill="x", padx=5, pady=2)

        prod_frame = tk.LabelFrame(sidebar, text="Production BOM")
        prod_frame.pack(fill="both", expand=True, padx=5, pady=4)

        self.prod_tree = ttk.Treeview(
            prod_frame,
            columns=("ref", "val", "unit"),
            show="headings",
            height=10,
        )
        self.prod_tree.heading("ref", text="Ref")
        self.prod_tree.heading("val", text="Value")
        self.prod_tree.heading("unit", text="Unit")
        self.prod_tree.column("ref", width=70, anchor="w")
        self.prod_tree.column("val", width=80, anchor="w")
        self.prod_tree.column("unit", width=60, anchor="w")
        self.prod_tree.pack(fill="both", expand=True)
        self.prod_tree.bind("<<TreeviewSelect>>", self.on_prod_tree_select)

        tk.Button(prod_frame, text="Refresh View",
                  command=self.refresh_production_tree).pack(fill="x", pady=2)

        details_frame = tk.LabelFrame(sidebar, text="Component Details")
        details_frame.pack(fill="x", padx=5, pady=4)

        tk.Label(details_frame, text="Tolerance:").grid(row=0, column=0, sticky="w")
        tk.Label(details_frame, textvariable=self.detail_tol_var).grid(row=0, column=1, sticky="w")

        tk.Label(details_frame, text="Size (EIA):").grid(row=1, column=0, sticky="w")
        tk.Label(details_frame, textvariable=self.detail_size_var).grid(row=1, column=1, sticky="w")

        tk.Label(details_frame, text="Rating:").grid(row=2, column=0, sticky="w")
        tk.Label(details_frame, textvariable=self.detail_rate_var).grid(row=2, column=1, sticky="w")

    # ---------------------- scale ----------------------------

    def set_scale(self, box_scale, xy_scale):
        self.box_scale = box_scale
        self.scale_factor = xy_scale

        if self.raw_xy_data:
            for ref, raw in self.raw_xy_data.items():
                if ref not in self.xy_data:
                    self.xy_data[ref] = {
                        "ref": ref,
                        "value": "",
                        "unit": "",
                        "angle": raw.get("angle", 0),
                        "comp_type": self.detect_type(ref),
                        "nc": False,
                    }
                self.xy_data[ref]["x"] = raw["x"] * self.scale_factor
                self.xy_data[ref]["y"] = raw["y"] * self.scale_factor

        self.redraw()

    # ---------------------- helpers --------------------------

    def auto_default_unit(self, ref, unit):
        if unit not in ["", None]:
            return unit
        r = ref.upper()
        if r.startswith("C"):
            return "nF"
        if r.startswith("R"):
            return "Ohms"
        if r.startswith("L"):
            return "nH"
        return ""

    def detect_type(self, ref):
        r = ref.upper()
        if r.startswith("C"):
            return "Capacitor"
        if r.startswith("R"):
            return "Resistor"
        if r.startswith("L"):
            return "Inductor"
        return "Unknown"

    # ---------------------- XY -------------------------------

    def load_xy_file(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not fp:
            return

        data = {}
        raw_data = {}
        try:
            with open(fp, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ref = (
                        row.get("ReferenceID")
                        or row.get("Reference")
                        or row.get("Ref")
                        or row.get("Designator")
                        or ""
                    ).strip()
                    if not ref:
                        continue
                    try:
                        x_raw = float((row.get("X") or "").strip())
                        y_raw = float((row.get("Y") or "").strip())
                        angle = float((row.get("Angle") or "0").strip())
                    except ValueError:
                        continue

                    raw_data[ref] = {"x": x_raw, "y": y_raw, "angle": angle}

                    x = x_raw * self.scale_factor
                    y = y_raw * self.scale_factor

                    data[ref] = {
                        "ref": ref,
                        "x": x,
                        "y": y,
                        "angle": angle,
                        "value": "",
                        "unit": "",
                        "comp_type": self.detect_type(ref),
                        "nc": False,
                    }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load XY file:\n{e}")
            return

        self.raw_xy_data = raw_data
        self.xy_data = data
        messagebox.showinfo("Loaded", "XY file loaded.")
        self.redraw()

    # ---------------------- production BOM -------------------

    def load_production_bom(self):
        fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not fp:
            return
        try:
            self.production_bom = ProductionBOM.from_xlsx(fp, self.auto_default_unit)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        for ref, info in self.xy_data.items():
            entry = self.production_bom.entries.get(ref)
            if entry:
                info["value"] = entry["value"]
                info["unit"] = entry["unit"]

        messagebox.showinfo("Loaded", "Production BOM applied.")
        self.refresh_production_tree()
        self.redraw()

    def refresh_production_tree(self):
        for item in self.prod_tree.get_children():
            self.prod_tree.delete(item)
        if not self.production_bom:
            return
        for ref in sorted(self.production_bom.entries.keys()):
            d = self.production_bom.entries[ref]
            self.prod_tree.insert(
                "",
                "end",
                values=(ref, d.get("value", ""), d.get("unit", "")),
            )

    def export_production_bom(self):
        if not self.production_bom:
            messagebox.showerror("Error", "No production BOM loaded.")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Export Production BOM",
        )
        if not save_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.append(["Reference Designator", "Value", "Unit"])
        for ref, d in self.production_bom.entries.items():
            ws.append([ref, d.get("value", ""), d.get("unit", "")])
        wb.save(save_path)
        messagebox.showinfo("Saved", f"Production BOM exported:\n{save_path}")

    # ---------------------- board versions -------------------

    def _get_raw_coords_for_board(self):
        if self.raw_xy_data:
            return {
                ref: {
                    "x": raw["x"],
                    "y": raw["y"],
                    "angle": raw.get("angle", 0),
                }
                for ref, raw in self.raw_xy_data.items()
            }
        data = {}
        for ref, info in self.xy_data.items():
            data[ref] = {
                "x": info["x"] / self.scale_factor,
                "y": info["y"] / self.scale_factor,
                "angle": info.get("angle", 0),
            }
        return data

    def create_board(self):
        if not self.xy_data:
            messagebox.showerror("Error", "Load an XY file first.")
            return None

        board_name = simpledialog.askstring("Create Board", "Board name:")
        if not board_name:
            return None

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Create Board",
            initialfile=f"{board_name}.xlsx",
        )
        if not save_path:
            return None

        board = Board(save_path)
        raw_coords = self._get_raw_coords_for_board()
        board.xy = raw_coords
        board.versions = []
        board.version_names = []
        board.version_meta = []
        try:
            board.save_to_xlsx()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create board:\n{e}")
            return None

        self.current_board = board
        self.current_board_version = None

        messagebox.showinfo("Created", f"Board created: {board.name}")
        return board

    def load_board_xlsx(self):
        fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not fp:
            return
        try:
            board = Board.load_from_xlsx(fp)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load board:\n{e}")
            return

        self.current_board = board
        self.current_board_version = 0 if board.versions else None

        self.raw_xy_data = {}
        self.xy_data = {}
        for ref, coords in board.xy.items():
            self.raw_xy_data[ref] = {
                "x": coords["x"],
                "y": coords["y"],
                "angle": coords["angle"],
            }
            self.xy_data[ref] = {
                "ref": ref,
                "x": coords["x"] * self.scale_factor,
                "y": coords["y"] * self.scale_factor,
                "angle": coords["angle"],
                "value": "",
                "unit": "",
                "comp_type": self.detect_type(ref),
                "nc": False,
            }

        if self.current_board_version is not None:
            ver = self.current_board.versions[self.current_board_version]
            for ref, info in self.xy_data.items():
                if ref in ver:
                    info["value"] = ver[ref]["value"]
                    info["unit"] = ver[ref]["unit"]

        messagebox.showinfo(
            "Loaded",
            f"Board loaded: {board.name}\nVersions: {', '.join(board.version_names) or 'none'}"
        )
        self.redraw()

    def save_new_board_version(self):
        if not self.current_board:
            create = messagebox.askyesno(
                "Create Board",
                "No board loaded. Create a new board?"
            )
            if create:
                if not self.create_board():
                    return
            else:
                return

        notes = simpledialog.askstring(
            "Version Notes",
            "Notes for this tuning version (optional):"
        )
        notes = notes or ""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        self.current_board.append_version_from_xy(self.xy_data, timestamp, notes)
        try:
            self.current_board.save_to_xlsx()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save board:\n{e}")
            return

        messagebox.showinfo(
            "Saved",
            f"New version saved as {self.current_board.version_names[-1]} in {self.current_board.name}"
        )

    def compare_board_versions(self):
        if not self.current_board or len(self.current_board.versions) < 2:
            messagebox.showerror("Error", "Load a board with at least two versions.")
            return

        win = tk.Toplevel()
        win.title("Compare Board Versions")

        names = self.current_board.version_names

        tk.Label(win, text="Select Version A:").pack()
        selA = tk.StringVar()
        comboA = ttk.Combobox(win, textvariable=selA, values=names, width=20)
        comboA.pack()
        comboA.current(0)

        tk.Label(win, text="Select Version B:").pack()
        selB = tk.StringVar()
        comboB = ttk.Combobox(win, textvariable=selB, values=names, width=20)
        comboB.pack()
        comboB.current(1)

        def do_compare():
            idxA = comboA.current()
            idxB = comboB.current()
            bomA = self.current_board.versions[idxA]
            bomB = self.current_board.versions[idxB]
            win.destroy()
            self.show_board_version_diff(names[idxA], names[idxB], bomA, bomB)

        tk.Button(win, text="Compare", command=do_compare).pack(pady=10)

    def show_board_version_diff(self, nameA, nameB, bomA, bomB):
        win = tk.Toplevel()
        win.title(f"Board Version Differences: {nameA} vs {nameB}")

        tree = ttk.Treeview(
            win,
            columns=("ref", "A_val", "B_val"),
            show="headings",
        )
        tree.pack(fill="both", expand=True)

        tree.heading("ref", text="Ref")
        tree.heading("A_val", text=nameA)
        tree.heading("B_val", text=nameB)

        refs = sorted(set(bomA.keys()) | set(bomB.keys()))
        for ref in refs:
            A_entry = bomA.get(ref, {})
            B_entry = bomB.get(ref, {})
            A_val = Board._join_value_unit(A_entry.get("value", ""), A_entry.get("unit", ""))
            B_val = Board._join_value_unit(B_entry.get("value", ""), B_entry.get("unit", ""))
            if A_val != B_val:
                tree.insert("", "end", values=(ref, A_val, B_val))

    # ---------------------- detail panel hooks ---------------

    def on_prod_tree_select(self, event):
        sel = self.prod_tree.selection()
        if not sel or not self.production_bom:
            return
        ref = self.prod_tree.item(sel[0])["values"][0]
        self.update_details(ref)

    def on_component_clicked(self, ref):
        self.update_details(ref)

    def update_details(self, ref):
        if not self.production_bom:
            self.detail_tol_var.set("")
            self.detail_size_var.set("")
            self.detail_rate_var.set("")
            return
        entry = self.production_bom.entries.get(ref, {})
        self.detail_tol_var.set(entry.get("tolerance", ""))
        self.detail_size_var.set(entry.get("size_eia", ""))
        self.detail_rate_var.set(entry.get("rating", ""))

    # ---------------------- redraw ---------------------------

    def redraw(self):
        self.canvas.delete("all")
        if not self.xy_data:
            return

        xs = [info["x"] for info in self.xy_data.values()]
        ys = [info["y"] for info in self.xy_data.values()]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)

        canvas_w = int(self.canvas.winfo_width() or 1500)
        canvas_h = int(self.canvas.winfo_height() or 900)

        width_span = max_x - min_x if max_x != min_x else 1
        height_span = max_y - min_y if max_y != min_y else 1

        offset_x = (canvas_w - width_span) / 2 - min_x
        offset_y = (canvas_h - height_span) / 2 - min_y

        for ref, info in self.xy_data.items():
            val = info.get("value", "")
            unit = info.get("unit", "")
            angle = info.get("angle", 0)
            nc_flag = info.get("nc", False)

            if nc_flag:
                val = ""
                unit = ""

            if val and not unit:
                unit = self.auto_default_unit(ref, unit)
                info["unit"] = unit

            highlight = None
            if val in ["", None]:
                highlight = "missing"
            else:
                if self.production_bom and not nc_flag:
                    entry = self.production_bom.entries.get(ref)
                    if entry:
                        pval = entry.get("value", "")
                        punit = entry.get("unit", "")
                        if pval or punit:
                            if not values_match(val, unit, pval, punit):
                                highlight = "mismatch"

            draw_x = info["x"] + offset_x
            draw_y = info["y"] + offset_y

            ComponentBox(
                self.canvas,
                ref,
                draw_x,
                draw_y,
                angle,
                comp_type=info.get("comp_type", "Unknown"),
                value=val,
                unit=unit,
                highlight=highlight,
                box_scale=self.box_scale,
                nc=nc_flag,
                on_left_click=self.on_component_clicked,
            )


# ============================================================
# Run
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    root.state("zoomed")
    app = LayoutAppV2(root)
    root.mainloop()
